# -*- coding: utf-8 -*-

from datetime import date
from io import BytesIO

from django.shortcuts import redirect
from django.shortcuts import render

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.http import FileResponse
from django.shortcuts import redirect, render


from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import Image
from reportlab.platypus import Paragraph
from reportlab.platypus import SimpleDocTemplate
from reportlab.platypus import Spacer
from reportlab.platypus import Table
from reportlab.platypus import TableStyle

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side
from openpyxl.worksheet.table import Table as ExcelTable
from openpyxl.worksheet.table import TableStyleInfo

from PIL import Image as PILImage
from PIL import ImageOps

from accounts.models import UserProfile
from docs.models import Document
from trips.models import Trip
from vehicles.models import Vehicle


@login_required
def startseite(request):

    planungen = Document.objects.filter(
        category="planung",
        active=True
    ).order_by(
        "-uploaded_at"
    )

    return render(
        request,
        "dashboard/startseite.html",
        {
            "planungen": planungen
        }
    )


@login_required
def home(request, year=None):

    today = date.today()

    base_year = today.year

    if year is None:

        current_year = base_year

    else:

        current_year = year

    profile = UserProfile.objects.filter(
        user=request.user
    ).first()

    if profile is None:

        return redirect("/login/")

    # =================================================
    # JAHRESNAVIGATION
    # =================================================

    available_years = set()

    # Aktuelles Jahr immer anzeigen

    available_years.add(
        base_year
    )

    # Vergangene Jahre nur anzeigen,
    # wenn in diesem Jahr Fahrten vorhanden sind

    trip_years = (
        Trip.objects
        .filter(
            user_profile=profile,
            date__lt=date(
                base_year,
                1,
                1
            )
        )
        .dates(
            "date",
            "year",
            order="DESC"
        )
    )

    for year_date in trip_years:

        available_years.add(
            year_date.year
        )

    # Ab 1. Dezember zusätzlich
    # das kommende Jahr anzeigen

    if today.month >= 12:

        available_years.add(
            base_year + 1
        )

    available_years = sorted(
        available_years,
        reverse=True
    )

    # =================================================
    # FAHRTEN
    # =================================================

    trips = Trip.objects.filter(
        user_profile=profile,
        date__year=current_year
    ).order_by(
        "-date"
    )

    recent_trips = trips

    total_hours = sum(
        trip.hours
        for trip in trips
    )

    # =================================================
    # SOLLSTUNDEN
    # =================================================

    if profile.category:

        target_hours = (
            profile.category.yearly_target_hours
        )

    else:

        target_hours = 0

    # =================================================
    # FORTSCHRITT
    # =================================================

    percent = 0

    if target_hours > 0:

        percent = round(
            (
                total_hours
                /
                target_hours
            )
            * 100,
            1
        )

    remaining = max(
        target_hours - total_hours,
        0
    )

    # =================================================
    # STATISTIK
    # =================================================

    trip_count = trips.count()

    vehicle_count = (
        trips
        .exclude(
            vehicle=None
        )
        .values(
            "vehicle"
        )
        .distinct()
        .count()
    )

    # =================================================
    # FAHRZEUG-REFRESH
    # =================================================

    vehicle_refresh = []

    vehicles = Vehicle.objects.filter(
        active=True
    )

    for vehicle in vehicles:

        last_trip = (
            Trip.objects
            .filter(
                vehicle=vehicle
            )
            .order_by(
                "-date"
            )
            .first()
        )

        if last_trip:

            days_since = (
                date.today()
                -
                last_trip.date
            ).days

            if days_since > 180:

                status = "Fällig"

                status_icon = (
                    "images/signal_faellig.PNG"
                )

            elif days_since > 90:

                status = "Warnung"

                status_icon = (
                    "images/signal_warnung.PNG"
                )

            else:

                status = "Aktuell"

                status_icon = (
                    "images/signal_ok.PNG"
                )

            vehicle_refresh.append({

                "vehicle": vehicle,

                "last_trip":
                    last_trip.date,

                "days_since":
                    days_since,

                "status":
                    status,

                "status_icon":
                    status_icon,

            })

    # =================================================
    # DASHBOARD
    # =================================================

    return render(

        request,

        "dashboard/home.html",

        {

            "profile":
                profile,

            "employee":
                profile,

            "base_year":
                base_year,

            "current_year":
                current_year,

            "available_years":
                available_years,

            "total_hours":
                total_hours,

            "target_hours":
                target_hours,

            "percent":
                percent,

            "remaining":
                remaining,

            "recent_trips":
                recent_trips,

            "employee_stats":
                [],

            "vehicle_refresh":
                vehicle_refresh,

            "trip_count":
                trip_count,

            "vehicle_count":
                vehicle_count,

        }

    )


# =====================================================
# PDF JAHRESAUSWERTUNG
# =====================================================

@login_required
def home_pdf(request, year):

    profile = UserProfile.objects.filter(
        user=request.user
    ).first()

    if profile is None:

        return redirect("/login/")

    # =================================================
    # FAHRTEN DES GEWÄHLTEN JAHRES
    # =================================================

    trips = (
        Trip.objects
        .filter(
            user_profile=profile,
            date__year=year
        )
        .order_by(
            "date"
        )
    )

    total_hours = sum(
        trip.hours
        for trip in trips
    )

    # =================================================
    # SOLLSTUNDEN
    # =================================================

    if profile.category:

        target_hours = (
            profile.category.yearly_target_hours
        )

    else:

        target_hours = 0

    # =================================================
    # RESTSTUNDEN
    # =================================================

    remaining = max(
        target_hours - total_hours,
        0
    )

    # =================================================
    # PROZENT
    # =================================================

    percent = 0

    if target_hours > 0:

        percent = round(
            (
                total_hours
                /
                target_hours
            )
            * 100,
            1
        )

    # Django Decimal für grafische Berechnungen
    # sicher in float umwandeln

    percent_float = float(
        percent
    )

    # =================================================
    # PDF BUFFER
    # =================================================

    buffer = BytesIO()

    document = SimpleDocTemplate(

        buffer,

        pagesize=A4,

        rightMargin=15 * mm,

        leftMargin=15 * mm,

        topMargin=18 * mm,

        bottomMargin=18 * mm,

        title=(
            f"Fahrpraxis {year}"
        ),

        author=(
            "Semita Fahrpraxis Manager"
        ),

    )

    styles = (
        getSampleStyleSheet()
    )

    # =================================================
    # FARBEN
    # =================================================

    semita_dark = colors.HexColor(
        "#0A2740"
    )

    semita_blue = colors.HexColor(
        "#244C82"
    )

    very_light = colors.HexColor(
        "#F5F7FB"
    )

    border = colors.HexColor(
        "#D9DFE5"
    )

    grey = colors.HexColor(
        "#666666"
    )

    green = colors.HexColor(
        "#2E7D32"
    )

    orange = colors.HexColor(
        "#EF6C00"
    )

    red = colors.HexColor(
        "#C62828"
    )

    white = colors.white

    # =================================================
    # TEXTSTYLES
    # =================================================

    title_style = ParagraphStyle(

        "PDFTitle",

        parent=styles["Title"],

        fontName="Helvetica-Bold",

        fontSize=23,

        leading=27,

        textColor=semita_blue,

        spaceAfter=4,

    )

    subtitle_style = ParagraphStyle(

        "PDFSubtitle",

        parent=styles["Normal"],

        fontName="Helvetica",

        fontSize=10,

        leading=13,

        textColor=grey,

    )

    section_style = ParagraphStyle(

        "PDFSection",

        parent=styles["Heading2"],

        fontName="Helvetica-Bold",

        fontSize=13,

        leading=16,

        textColor=semita_blue,

        spaceBefore=8,

        spaceAfter=8,

    )

    normal_style = ParagraphStyle(

        "PDFNormal",

        parent=styles["Normal"],

        fontName="Helvetica",

        fontSize=9,

        leading=12,

        textColor=colors.HexColor(
            "#333333"
        ),

    )

    white_small_style = ParagraphStyle(

        "PDFWhiteSmall",

        parent=styles["Normal"],

        fontName="Helvetica-Bold",

        fontSize=7,

        leading=9,

        textColor=white,

    )

    stat_title_style = ParagraphStyle(

        "PDFStatTitle",

        parent=styles["Normal"],

        fontName="Helvetica",

        fontSize=8,

        leading=10,

        textColor=grey,

    )

    stat_value_style = ParagraphStyle(

        "PDFStatValue",

        parent=styles["Normal"],

        fontName="Helvetica-Bold",

        fontSize=18,

        leading=21,

        textColor=semita_blue,

    )

    legal_style = ParagraphStyle(

        "PDFLegalText",

        parent=normal_style,

        fontName="Helvetica",

        fontSize=8.5,

        leading=11,

        textColor=colors.HexColor(
            "#333333"
        ),

    )

    # =================================================
    # STORY
    # =================================================

    story = []

    # =================================================
    # LOGO
    # =================================================

    logo_path = (

        settings.BASE_DIR
        / "static"
        / "images"
        / "semita.png"

    )

    if logo_path.exists():

        logo = Image(

            str(logo_path),

            width=45 * mm,

            height=15 * mm,

            kind="proportional",

        )

        logo.hAlign = "LEFT"

    else:

        logo = Paragraph(

            "<b>SEMITA</b>",

            ParagraphStyle(

                "LogoFallback",

                parent=title_style,

                fontSize=20,

                textColor=white,

            )

        )

    # =================================================
    # KOPFZEILE
    # =================================================

    header_right = Paragraph(

        f"<b>FAHRPRAXIS</b><br/>"
        f"Jahresauswertung {year}",

        ParagraphStyle(

            "HeaderRight",

            parent=styles["Normal"],

            fontName="Helvetica-Bold",

            fontSize=9,

            leading=13,

            alignment=2,

            textColor=white,

        )

    )

    header_table = Table(

        [

            [

                logo,

                header_right,

            ]

        ],

        colWidths=[

            90 * mm,

            90 * mm,

        ],

        rowHeights=[

            24 * mm

        ],

    )

    header_table.setStyle(

        TableStyle([

            (

                "BACKGROUND",

                (0, 0),

                (-1, -1),

                semita_dark,

            ),

            (

                "VALIGN",

                (0, 0),

                (-1, -1),

                "MIDDLE",

            ),

            (

                "ALIGN",

                (1, 0),

                (1, 0),

                "RIGHT",

            ),

            (

                "LEFTPADDING",

                (0, 0),

                (-1, -1),

                8,

            ),

            (

                "RIGHTPADDING",

                (0, 0),

                (-1, -1),

                8,

            ),

            (

                "TOPPADDING",

                (0, 0),

                (-1, -1),

                5,

            ),

            (

                "BOTTOMPADDING",

                (0, 0),

                (-1, -1),

                5,

            ),

        ])

    )

    story.append(
        header_table
    )

    story.append(
        Spacer(
            1,
            8 * mm
        )
    )

    # =================================================
    # TITEL
    # =================================================

    story.append(

        Paragraph(

            f"Fahrpraxis {year}",

            title_style

        )

    )

    story.append(

        Paragraph(

            "Jahresauswertung der "
            "Mindestfahrpraxis",

            subtitle_style

        )

    )

    story.append(

        Spacer(
            1,
            5 * mm
        )

    )

    # =================================================
    # MITARBEITER
    # =================================================

    story.append(

        Paragraph(

            "Mitarbeiter",

            section_style

        )

    )

    full_name = (

        profile.user.get_full_name()

        or profile.user.email

    )

    category = (

        str(profile.category)

        if profile.category

        else "Keine Kategorie"

    )

    employee_data = [

        [

            Paragraph(
                "<b>Name</b>",
                normal_style
            ),

            Paragraph(
                full_name,
                normal_style
            ),

            Paragraph(
                "<b>Kategorie</b>",
                normal_style
            ),

            Paragraph(
                category,
                normal_style
            ),

        ],

        [

            Paragraph(
                "<b>E-Mail</b>",
                normal_style
            ),

            Paragraph(
                profile.user.email or "-",
                normal_style
            ),

            Paragraph(
                "<b>Jahr</b>",
                normal_style
            ),

            Paragraph(
                str(year),
                normal_style
            ),

        ],

    ]

    employee_table = Table(

        employee_data,

        colWidths=[

            25 * mm,

            65 * mm,

            25 * mm,

            65 * mm,

        ]

    )

    employee_table.setStyle(

        TableStyle([

            (

                "BACKGROUND",

                (0, 0),

                (-1, -1),

                very_light,

            ),

            (

                "BOX",

                (0, 0),

                (-1, -1),

                0.7,

                border,

            ),

            (

                "INNERGRID",

                (0, 0),

                (-1, -1),

                0.3,

                border,

            ),

            (

                "VALIGN",

                (0, 0),

                (-1, -1),

                "MIDDLE",

            ),

            (

                "TOPPADDING",

                (0, 0),

                (-1, -1),

                8,

            ),

            (

                "BOTTOMPADDING",

                (0, 0),

                (-1, -1),

                8,

            ),

            (

                "LEFTPADDING",

                (0, 0),

                (-1, -1),

                8,

            ),

            (

                "RIGHTPADDING",

                (0, 0),

                (-1, -1),

                8,

            ),

        ])

    )

    story.append(
        employee_table
    )

    # =================================================
    # RECHTLICHE GRUNDLAGEN
    # =================================================

    story.append(

        Spacer(
            1,
            5 * mm
        )

    )

    story.append(

        Paragraph(

            "Rechtliche Grundlagen",

            section_style

        )

    )

    legal_text = (

        "<b>Auszug: Verordnung des UVEK über die Zulassung "
        "zum Führen von Triebfahrzeugen der Eisenbahnen (VTE)</b>"
        "<br/>"

        "<b>4. Kapitel: Fahrpraxis &gt; Art. 34 Allgemeines</b>"
        "<br/><br/>"

        "<b>1</b>&nbsp;&nbsp;"
        "Die Fahrpraxis ist durch Tätigkeiten im Rahmen der "
        "Bescheinigung zu erwerben."

        "<br/><br/>"

        "<b>2</b>&nbsp;&nbsp;"
        "Lokführer und -führerinnen der Kategorien B60, B80, "
        "B100 und B sowie Strassenbahnführer und -führerinnen "
        "können die Hälfte der Fahrpraxis durch Pilotieren "
        "erwerben, wobei eine Pilotierstunde als halbe Fahrstunde "
        "zählt."

        "<br/><br/>"

        "<b>I-12470: 3.6. Fehlende Mindestfahrpraxis</b>"

        "<br/><br/>"

        "Kann die nötige Fahrpraxis gemäss VTE Art. 35 nicht "
        "nachgewiesen werden, muss vor dem selbstständigen Einsatz "
        "eine praktische Prüfung gemäss VTE Art. 37 absolviert "
        "werden. Die Prüfung ist bei einem PEX BAV zu bestellen."

    )

    legal_box = Table(

        [

            [

                Paragraph(
                    legal_text,
                    legal_style
                )

            ]

        ],

        colWidths=[

            180 * mm

        ]

    )

    legal_box.setStyle(

        TableStyle([

            (

                "BACKGROUND",

                (0, 0),

                (-1, -1),

                very_light,

            ),

            (

                "BOX",

                (0, 0),

                (-1, -1),

                0.7,

                border,

            ),

            (

                "LEFTPADDING",

                (0, 0),

                (-1, -1),

                10,

            ),

            (

                "RIGHTPADDING",

                (0, 0),

                (-1, -1),

                10,

            ),

            (

                "TOPPADDING",

                (0, 0),

                (-1, -1),

                8,

            ),

            (

                "BOTTOMPADDING",

                (0, 0),

                (-1, -1),

                8,

            ),

        ])

    )

    story.append(
        legal_box
    )

    # =================================================
    # KENNZAHLEN
    # =================================================

    story.append(

        Paragraph(

            "Fahrpraxis",

            section_style

        )

    )

    stat_data = [

        [

            Paragraph(
                "SOLLSTUNDEN",
                stat_title_style
            ),

            Paragraph(
                "ISTSTUNDEN",
                stat_title_style
            ),

            Paragraph(
                "RESTSTUNDEN",
                stat_title_style
            ),

            Paragraph(
                "FORTSCHRITT",
                stat_title_style
            ),

        ],

        [

            Paragraph(
                f"{target_hours} h",
                stat_value_style
            ),

            Paragraph(
                f"{total_hours} h",
                stat_value_style
            ),

            Paragraph(
                f"{remaining} h",
                stat_value_style
            ),

            Paragraph(
                f"{percent} %",
                stat_value_style
            ),

        ],

    ]

    stat_table = Table(

        stat_data,

        colWidths=[

            45 * mm,

            45 * mm,

            45 * mm,

            45 * mm,

        ],

        rowHeights=[

            9 * mm,

            15 * mm,

        ],

    )

    stat_table.setStyle(

        TableStyle([

            (

                "BACKGROUND",

                (0, 0),

                (-1, -1),

                very_light,

            ),

            (

                "BOX",

                (0, 0),

                (-1, -1),

                0.7,

                border,

            ),

            (

                "INNERGRID",

                (0, 0),

                (-1, -1),

                0.3,

                border,

            ),

            (

                "VALIGN",

                (0, 0),

                (-1, -1),

                "MIDDLE",

            ),

            (

                "LEFTPADDING",

                (0, 0),

                (-1, -1),

                8,

            ),

            (

                "RIGHTPADDING",

                (0, 0),

                (-1, -1),

                8,

            ),

        ])

    )

    story.append(
        stat_table
    )

    # =================================================
    # FORTSCHRITTSBALKEN
    # =================================================

    story.append(

        Spacer(
            1,
            5 * mm
        )

    )

    progress_value = min(
        max(percent_float, 0.0),
        100.0
    )

    if percent_float >= 100:

        progress_color = green

    elif percent_float >= 75:

        progress_color = semita_blue

    elif percent_float >= 50:

        progress_color = orange

    else:

        progress_color = red

    total_width = 180.0 * mm

    filled_width = (
        total_width
        *
        progress_value
        /
        100.0
    )

    empty_width = (
        total_width
        -
        filled_width
    )

    if filled_width > 0 and empty_width > 0:

        progress_table = Table(

            [
                [
                    "",
                    ""
                ]
            ],

            colWidths=[

                filled_width,

                empty_width,

            ],

            rowHeights=[

                7 * mm

            ],

        )

        progress_table.setStyle(

            TableStyle([

                (

                    "BACKGROUND",

                    (0, 0),

                    (0, 0),

                    progress_color,

                ),

                (

                    "BACKGROUND",

                    (1, 0),

                    (1, 0),

                    colors.HexColor(
                        "#D9DFE5"
                    ),

                ),

                (

                    "LEFTPADDING",

                    (0, 0),

                    (-1, -1),

                    0,

                ),

                (

                    "RIGHTPADDING",

                    (0, 0),

                    (-1, -1),

                    0,

                ),

                (

                    "TOPPADDING",

                    (0, 0),

                    (-1, -1),

                    0,

                ),

                (

                    "BOTTOMPADDING",

                    (0, 0),

                    (-1, -1),

                    0,

                ),

            ])

        )

    elif filled_width >= total_width:

        progress_table = Table(

            [
                [
                    ""
                ]
            ],

            colWidths=[

                total_width

            ],

            rowHeights=[

                7 * mm

            ],

        )

        progress_table.setStyle(

            TableStyle([

                (

                    "BACKGROUND",

                    (0, 0),

                    (-1, -1),

                    progress_color,

                ),

                (

                    "LEFTPADDING",

                    (0, 0),

                    (-1, -1),

                    0,

                ),

                (

                    "RIGHTPADDING",

                    (0, 0),

                    (-1, -1),

                    0,

                ),

                (

                    "TOPPADDING",

                    (0, 0),

                    (-1, -1),

                    0,

                ),

                (

                    "BOTTOMPADDING",

                    (0, 0),

                    (-1, -1),

                    0,

                ),

            ])

        )

    else:

        progress_table = Table(

            [
                [
                    ""
                ]
            ],

            colWidths=[

                total_width

            ],

            rowHeights=[

                7 * mm

            ],

        )

        progress_table.setStyle(

            TableStyle([

                (

                    "BACKGROUND",

                    (0, 0),

                    (-1, -1),

                    colors.HexColor(
                        "#D9DFE5"
                    ),

                ),

                (

                    "LEFTPADDING",

                    (0, 0),

                    (-1, -1),

                    0,

                ),

                (

                    "RIGHTPADDING",

                    (0, 0),

                    (-1, -1),

                    0,

                ),

                (

                    "TOPPADDING",

                    (0, 0),

                    (-1, -1),

                    0,

                ),

                (

                    "BOTTOMPADDING",

                    (0, 0),

                    (-1, -1),

                    0,

                ),

            ])

        )

    story.append(
        progress_table
    )

    # =================================================
    # STATUS
    # =================================================

    if remaining <= 0:

        status_text = (
            "Ô£ô Mindestfahrpraxis erfüllt"
        )

        status_color = green

    else:

        status_text = (
            f"{remaining} Stunden bis zur "
            "Mindestfahrpraxis"
        )

        status_color = semita_blue

    story.append(

        Spacer(
            1,
            3 * mm
        )

    )

    story.append(

        Paragraph(

            status_text,

            ParagraphStyle(

                "Status",

                parent=normal_style,

                fontName="Helvetica-Bold",

                fontSize=9,

                textColor=status_color,

            )

        )

    )

    # =================================================
    # FAHRTEN
    # =================================================

    story.append(

        Paragraph(

            f"Fahrten {year}",

            section_style

        )

    )

    trip_data = [

        [

            Paragraph(
                "<b>Datum</b>",
                white_small_style
            ),

            Paragraph(
                "<b>Art</b>",
                white_small_style
            ),

            Paragraph(
                "<b>Zug</b>",
                white_small_style
            ),

            Paragraph(
                "<b>Von</b>",
                white_small_style
            ),

            Paragraph(
                "<b>Nach</b>",
                white_small_style
            ),

            Paragraph(
                "<b>Fahrzeug</b>",
                white_small_style
            ),

            Paragraph(
                "<b>Std.</b>",
                white_small_style
            ),

        ]

    ]

    for trip in trips:

        if trip.traffic_type == "zug":

            traffic = "Zug"

        elif trip.traffic_type == "rangieren":

            traffic = "Rangierbewegung"

        else:

            traffic = str(
                trip.traffic_type
            )

        trip_data.append(

            [

                trip.date.strftime(
                    "%d.%m.%Y"
                ),

                traffic,

                trip.train_number or "-",

                (
                    str(trip.from_station)
                    if trip.from_station
                    else "-"
                ),

                (
                    str(trip.to_station)
                    if trip.to_station
                    else "-"
                ),

                (
                    str(trip.vehicle)
                    if trip.vehicle
                    else "-"
                ),

                str(trip.hours),

            ]

        )

    if len(trip_data) == 1:

        trip_data.append(

            [

                "-",

                "-",

                "-",

                "-",

                "-",

                "-",

                "Keine Fahrten",

            ]

        )

    trip_table = Table(

        trip_data,

        repeatRows=1,

        colWidths=[

            22 * mm,

            28 * mm,

            20 * mm,

            30 * mm,

            30 * mm,

            35 * mm,

            15 * mm,

        ]

    )

    trip_table.setStyle(

        TableStyle([

            (

                "BACKGROUND",

                (0, 0),

                (-1, 0),

                semita_dark,

            ),

            (

                "TEXTCOLOR",

                (0, 0),

                (-1, 0),

                white,

            ),

            (

                "FONTNAME",

                (0, 0),

                (-1, 0),

                "Helvetica-Bold",

            ),

            (

                "FONTSIZE",

                (0, 0),

                (-1, -1),

                7,

            ),

            (

                "GRID",

                (0, 0),

                (-1, -1),

                0.3,

                border,

            ),

            (

                "ROWBACKGROUNDS",

                (0, 1),

                (-1, -1),

                [

                    white,

                    very_light,

                ],

            ),

            (

                "VALIGN",

                (0, 0),

                (-1, -1),

                "MIDDLE",

            ),

            (

                "TOPPADDING",

                (0, 0),

                (-1, -1),

                6,

            ),

            (

                "BOTTOMPADDING",

                (0, 0),

                (-1, -1),

                6,

            ),

            (

                "LEFTPADDING",

                (0, 0),

                (-1, -1),

                5,

            ),

            (

                "RIGHTPADDING",

                (0, 0),

                (-1, -1),

                5,

            ),

        ])

    )

    story.append(
        trip_table
    )

    # =================================================
    # FOOTER
    # =================================================

    def add_footer(
        canvas,
        doc
    ):

        canvas.saveState()

        width, height = A4

        canvas.setStrokeColor(
            border
        )

        canvas.line(

            15 * mm,

            12 * mm,

            width - 15 * mm,

            12 * mm,

        )

        canvas.setFont(
            "Helvetica",
            7
        )

        canvas.setFillColor(
            grey
        )

        canvas.drawString(

            15 * mm,

            7 * mm,

            "Semita Fahrpraxis Manager"

        )

        canvas.drawRightString(

            width - 15 * mm,

            7 * mm,

            f"Seite {doc.page}"

        )

        canvas.restoreState()

    # =================================================
    # PDF ERSTELLEN
    # =================================================

    document.build(

        story,

        onFirstPage=add_footer,

        onLaterPages=add_footer,

    )

    buffer.seek(0)

    filename = (

        f"Fahrpraxis_"

        f"{profile.user.last_name}_"

        f"{year}.pdf"

    )

    return FileResponse(

        buffer,

        as_attachment=True,

        filename=filename,

        content_type="application/pdf",

    )


# =====================================================
# EXCEL JAHRESAUSWERTUNG
# =====================================================

@login_required
def home_excel(request, year):

    profile = UserProfile.objects.filter(
        user=request.user
    ).first()

    if profile is None:
        return redirect("/login/")

    # =================================================
    # FAHRTEN
    # =================================================

    trips = (
        Trip.objects
        .filter(
            user_profile=profile,
            date__year=year
        )
        .order_by("date")
    )

    total_hours = sum(
        trip.hours
        for trip in trips
    )

    # =================================================
    # SOLLSTUNDEN
    # =================================================

    if profile.category:

        target_hours = (
            profile.category.yearly_target_hours
        )

    else:

        target_hours = 0

    # =================================================
    # RESTSTUNDEN
    # =================================================

    remaining = max(
        target_hours - total_hours,
        0
    )

    # =================================================
    # FORTSCHRITT
    # =================================================

    if target_hours > 0:

        percent = round(
            (
                total_hours
                /
                target_hours
            ) * 100,
            1
        )

    else:

        percent = 0

    # =================================================
    # WORKBOOK
    # =================================================

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = (
        f"Fahrpraxis {year}"
    )

    # =================================================
    # FARBEN
    # =================================================

    semita_dark = "0A2740"

    semita_blue = "244C82"

    very_light = "F5F7FB"

    border_color = "D9DFE5"

    grey = "666666"

    white = "FFFFFF"

    green = "2E7D32"

    orange = "EF6C00"

    red = "C62828"

    thin = Side(
        style="thin",
        color=border_color
    )

    # =================================================
    # SPALTENBREITEN
    # =================================================

    column_widths = {

        "A": 15,

        "B": 22,

        "C": 15,

        "D": 25,

        "E": 25,

        "F": 24,

        "G": 12,

    }

    for column, width in column_widths.items():

        worksheet.column_dimensions[
            column
        ].width = width

    # =================================================
    # KOPF
    # =================================================

    for row in worksheet["A1:G4"]:

        for cell in row:

            cell.fill = PatternFill(
                "solid",
                fgColor=semita_dark
            )

    worksheet.merge_cells(
        "A1:G4"
    )

    worksheet["A1"] = (
        f"FAHRPRAXIS\n"
        f"JAHRESAUSWERTUNG {year}"
    )

    worksheet["A1"].font = Font(
        name="Arial",
        size=20,
        bold=True,
        color=white
    )

    worksheet["A1"].alignment = Alignment(
        horizontal="right",
        vertical="center",
        wrap_text=True
    )

    # =================================================
    # LOGO
    #
    # WICHTIG:
    # Das Logo wird proportional skaliert.
    # Es wird NICHT mehr mit width und height
    # unabhängig voneinander verzerrt.
    # =================================================

    logo_path = (
        settings.BASE_DIR
        / "static"
        / "images"
        / "semita.png"
    )

    if logo_path.exists():

        try:

            # -----------------------------------------
            # Originalbild Öffnen
            # -----------------------------------------

            source_logo = PILImage.open(
                logo_path
            )

            # EXIF-Ausrichtung korrigieren
            source_logo = ImageOps.exif_transpose(
                source_logo
            )

            # -----------------------------------------
            # RGBA verwenden
            # -----------------------------------------

            source_logo = source_logo.convert(
                "RGBA"
            )

            # -----------------------------------------
            # Transparente Ränder entfernen
            # -----------------------------------------

            alpha = source_logo.getchannel(
                "A"
            )

            bbox = alpha.getbbox()

            if bbox:

                source_logo = source_logo.crop(
                    bbox
                )

            # -----------------------------------------
            # OriginalgrÖsse
            # -----------------------------------------

            original_width, original_height = (
                source_logo.size
            )

            # -----------------------------------------
            # Maximale GrÖsse im Excel-Kopf
            #
            # Das Seitenverhältnis bleibt erhalten.
            # -----------------------------------------

            max_logo_width = 120

            max_logo_height = 65

            scale = min(

                max_logo_width
                /
                original_width,

                max_logo_height
                /
                original_height

            )

            logo_width = max(
                1,
                int(
                    original_width
                    * scale
                )
            )

            logo_height = max(
                1,
                int(
                    original_height
                    * scale
                )
            )

            # -----------------------------------------
            # Proportional verkleinern
            # -----------------------------------------

            source_logo = source_logo.resize(

                (
                    logo_width,
                    logo_height
                ),

                PILImage.Resampling.LANCZOS

            )

            # -----------------------------------------
            # PNG in BytesIO schreiben
            #
            # NICHT das PIL-Objekt direkt an
            # openpyxl übergeben.
            # -----------------------------------------

            logo_buffer = BytesIO()

            source_logo.save(

                logo_buffer,

                format="PNG"

            )

            logo_buffer.seek(0)

            # -----------------------------------------
            # openpyxl bekommt den BytesIO-Stream
            # -----------------------------------------

            logo = ExcelImage(
                logo_buffer
            )

            # Sicherheitshalber nochmals die
            # berechnete proportionale GrÖsse setzen

            logo.width = logo_width

            logo.height = logo_height

            worksheet.add_image(
                logo,
                "A1"
            )

        except Exception:
            # Excel-Datei soll auch dann funktionieren,
            # wenn das Logo nicht geladen werden kann.
            pass

    # =================================================
    # TITEL
    # =================================================

    worksheet.merge_cells(
        "A6:G6"
    )

    worksheet["A6"] = (
        f"Fahrpraxis {year}"
    )

    worksheet["A6"].font = Font(
        name="Arial",
        size=20,
        bold=True,
        color=semita_blue
    )

    worksheet["A6"].alignment = Alignment(
        vertical="center"
    )

    worksheet.merge_cells(
        "A7:G7"
    )

    worksheet["A7"] = (
        "Jahresauswertung der "
        "Mindestfahrpraxis"
    )

    worksheet["A7"].font = Font(
        name="Arial",
        size=10,
        color=grey
    )

    # =================================================
    # MITARBEITER
    # =================================================

    worksheet.merge_cells(
        "A9:G9"
    )

    worksheet["A9"] = (
        "Mitarbeiter"
    )

    worksheet["A9"].font = Font(
        name="Arial",
        size=13,
        bold=True,
        color=semita_blue
    )

    full_name = (

        profile.user.get_full_name()

        or profile.user.email

    )

    category = (

        str(profile.category)

        if profile.category

        else "Keine Kategorie"

    )

    employee_rows = [

        (
            "Name",
            full_name,
            "Kategorie",
            category
        ),

        (
            "E-Mail",
            profile.user.email or "-",
            "Jahr",
            year
        ),

    ]

    for row_number, values in enumerate(
        employee_rows,
        10
    ):

        worksheet.cell(
            row_number,
            1,
            values[0]
        )

        worksheet.cell(
            row_number,
            2,
            values[1]
        )

        worksheet.cell(
            row_number,
            4,
            values[2]
        )

        worksheet.cell(
            row_number,
            5,
            values[3]
        )

        for column in range(
            1,
            7
        ):

            cell = worksheet.cell(
                row_number,
                column
            )

            cell.fill = PatternFill(
                "solid",
                fgColor=very_light
            )

            cell.border = Border(
                top=thin,
                bottom=thin,
                left=thin,
                right=thin
            )

            cell.alignment = Alignment(
                vertical="center"
            )

        for column in (
            1,
            4
        ):

            worksheet.cell(
                row_number,
                column
            ).font = Font(
                name="Arial",
                bold=True
            )

    # =================================================
    # RECHTLICHE GRUNDLAGEN
    # =================================================

    worksheet.merge_cells(
        "A13:G13"
    )

    worksheet["A13"] = (
        "Rechtliche Grundlagen"
    )

    worksheet["A13"].font = Font(
        name="Arial",
        size=13,
        bold=True,
        color=semita_blue
    )

    legal_text = (

        "Auszug: Verordnung des UVEK "
        "über die Zulassung zum Führen "
        "von Triebfahrzeugen der Eisenbahnen "
        "(VTE)\n"

        "4. Kapitel: Fahrpraxis > "
        "Art. 34 Allgemeines\n\n"

        "1  Die Fahrpraxis ist durch Tätigkeiten "
        "im Rahmen der Bescheinigung zu erwerben.\n\n"

        "2  Lokführer und -führerinnen der "
        "Kategorien B60, B80, B100 und B sowie "
        "Strassenbahnführer und -führerinnen "
        "können die Hälfte der Fahrpraxis "
        "durch Pilotieren erwerben, wobei eine "
        "Pilotierstunde als halbe Fahrstunde zählt.\n\n"

        "I-12470: 3.6. Fehlende Mindestfahrpraxis\n\n"

        "Kann die nötige Fahrpraxis gemäss "
        "VTE Art. 35 nicht nachgewiesen werden, "
        "muss vor dem selbstständigen Einsatz "
        "eine praktische Prüfung gemäss VTE "
        "Art. 37 absolviert werden. Die Prüfung "
        "ist bei einem PEX BAV zu bestellen."

    )

    worksheet.merge_cells(
        "A14:G20"
    )

    worksheet["A14"] = legal_text

    worksheet["A14"].font = Font(
        name="Arial",
        size=9,
        color="333333"
    )

    worksheet["A14"].alignment = Alignment(
        vertical="top",
        wrap_text=True
    )

    for row in worksheet[
        "A14:G20"
    ]:

        for cell in row:

            cell.fill = PatternFill(
                "solid",
                fgColor=very_light
            )

            cell.border = Border(
                top=thin,
                bottom=thin,
                left=thin,
                right=thin
            )

    # =================================================
    # FAHRPRAXIS KENNZAHLEN
    # =================================================

    worksheet.merge_cells(
        "A22:G22"
    )

    worksheet["A22"] = (
        "Fahrpraxis"
    )

    worksheet["A22"].font = Font(
        name="Arial",
        size=13,
        bold=True,
        color=semita_blue
    )

    statistics = [

        (
            "SOLLSTUNDEN",
            target_hours
        ),

        (
            "ISTSTUNDEN",
            total_hours
        ),

        (
            "RESTSTUNDEN",
            remaining
        ),

        (
            "FORTSCHRITT",
            f"{percent} %"
        ),

    ]

    statistic_columns = (
        1,
        3,
        5,
        7
    )

    for (
        statistic,
        column
    ) in zip(
        statistics,
        statistic_columns
    ):

        label, value = statistic

        worksheet.cell(
            23,
            column,
            label
        )

        worksheet.cell(
            23,
            column
        ).font = Font(
            name="Arial",
            size=9,
            color=grey
        )

        worksheet.cell(
            24,
            column,
            value
        )

        worksheet.cell(
            24,
            column
        ).font = Font(
            name="Arial",
            size=17,
            bold=True,
            color=semita_blue
        )

        for row_number in (
            23,
            24
        ):

            cell = worksheet.cell(
                row_number,
                column
            )

            cell.fill = PatternFill(
                "solid",
                fgColor=very_light
            )

            cell.border = Border(
                top=thin,
                bottom=thin,
                left=thin,
                right=thin
            )

    # =================================================
    # FORTSCHRITT
    # =================================================

    worksheet.merge_cells(
        "A26:G26"
    )

    worksheet["A26"] = (
        "Fortschritt"
    )

    worksheet["A26"].font = Font(
        name="Arial",
        bold=True,
        color=grey
    )

    progress = min(
        max(
            float(percent),
            0
        ),
        100
    )

    if percent >= 100:

        progress_color = green

    elif percent >= 75:

        progress_color = semita_blue

    elif percent >= 50:

        progress_color = orange

    else:

        progress_color = red

    worksheet.merge_cells(
        "A27:G27"
    )

    worksheet["A27"] = (
        f"{percent} %"
    )

    worksheet["A27"].font = Font(
        name="Arial",
        bold=True,
        color=white
    )

    worksheet["A27"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    worksheet["A27"].fill = PatternFill(
        "solid",
        fgColor=progress_color
    )

    worksheet.row_dimensions[
        27
    ].height = 22

    # =================================================
    # STATUS
    # =================================================

    worksheet.merge_cells(
        "A29:G29"
    )

    if remaining <= 0:

        worksheet["A29"] = (
            "Ô£ô Mindestfahrpraxis erfüllt"
        )

        status_color = green

    else:

        worksheet["A29"] = (
            f"{remaining} Stunden bis zur "
            "Mindestfahrpraxis"
        )

        status_color = semita_blue

    worksheet["A29"].font = Font(
        name="Arial",
        bold=True,
        color=status_color
    )

    # =================================================
    # FAHRTEN
    # =================================================

    worksheet.merge_cells(
        "A31:G31"
    )

    worksheet["A31"] = (
        f"Fahrten {year}"
    )

    worksheet["A31"].font = Font(
        name="Arial",
        size=13,
        bold=True,
        color=semita_blue
    )

    headers = [

        "Datum",

        "Art",

        "Zug",

        "Von",

        "Nach",

        "Fahrzeug",

        "Std."

    ]

    for column, header in enumerate(
        headers,
        1
    ):

        cell = worksheet.cell(
            32,
            column,
            header
        )

        cell.font = Font(
            name="Arial",
            bold=True,
            color=white
        )

        cell.fill = PatternFill(
            "solid",
            fgColor=semita_dark
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        cell.border = Border(
            top=thin,
            bottom=thin,
            left=thin,
            right=thin
        )

    # =================================================
    # FAHRTEN EINTRAGEN
    # =================================================

    for row_number, trip in enumerate(
        trips,
        33
    ):

        if trip.traffic_type == "zug":

            traffic = "Zug"

        elif trip.traffic_type == "rangieren":

            traffic = "Rangierbewegung"

        else:

            traffic = str(
                trip.traffic_type
            )

        values = [

            trip.date,

            traffic,

            trip.train_number
            or "-",

            str(trip.from_station)
            if trip.from_station
            else "-",

            str(trip.to_station)
            if trip.to_station
            else "-",

            (
                str(trip.vehicle)
                if trip.vehicle
                else "-"
            ),

            trip.hours

        ]

        for column, value in enumerate(
            values,
            1
        ):

            cell = worksheet.cell(
                row_number,
                column,
                value
            )

            cell.font = Font(
                name="Arial",
                size=10
            )

            cell.border = Border(
                top=thin,
                bottom=thin,
                left=thin,
                right=thin
            )

            cell.alignment = Alignment(
                vertical="center"
            )

            if row_number % 2:

                cell.fill = PatternFill(
                    "solid",
                    fgColor=very_light
                )

            else:

                cell.fill = PatternFill(
                    "solid",
                    fgColor=white
                )

        worksheet.cell(
            row_number,
            1
        ).number_format = (
            "DD.MM.YYYY"
        )

        worksheet.cell(
            row_number,
            7
        ).number_format = (
            "0.00"
        )

    # =================================================
    # LEERE FAHRTENLISTE
    # =================================================

    if trips.count() == 0:

        worksheet.merge_cells(
            "A33:G33"
        )

        worksheet["A33"] = (
            "Keine Fahrten vorhanden"
        )

        worksheet["A33"].font = Font(
            name="Arial",
            size=10,
            color=grey
        )

        worksheet["A33"].alignment = Alignment(
            horizontal="center"
        )

    # =================================================
    # EXCEL-TABELLE
    # =================================================

    last_row = max(
        33,
        32 + trips.count()
    )

    if trips.count() > 0:

        excel_table = ExcelTable(

            displayName="FahrtenTabelle",

            ref=(
                f"A32:G{last_row}"
            )

        )

        table_style = TableStyleInfo(

            name="TableStyleMedium2",

            showFirstColumn=False,

            showLastColumn=False,

            showRowStripes=True,

            showColumnStripes=False

        )

        excel_table.tableStyleInfo = (
            table_style
        )

        worksheet.add_table(
            excel_table
        )

    # =================================================
    # EINFRIEREN
    # =================================================

    worksheet.freeze_panes = (
        "A33"
    )

    # =================================================
    # FILTER
    # =================================================

    worksheet.auto_filter.ref = (
        f"A32:G{last_row}"
    )

    # =================================================
    # GITTERNETZ AUSBLENDEN
    # =================================================

    worksheet.sheet_view.showGridLines = (
        False
    )

    # =================================================
    # DRUCKEINSTELLUNGEN
    # =================================================

    worksheet.page_setup.orientation = (
        "landscape"
    )

    worksheet.page_setup.paperSize = (
        worksheet.PAPERSIZE_A4
    )

    worksheet.page_setup.fitToWidth = 1

    worksheet.page_setup.fitToHeight = 0

    worksheet.sheet_properties.pageSetUpPr.fitToPage = (
        True
    )

    worksheet.page_margins.left = 0.3

    worksheet.page_margins.right = 0.3

    worksheet.page_margins.top = 0.5

    worksheet.page_margins.bottom = 0.5

    # =================================================
    # FOOTER
    # =================================================

    worksheet.oddFooter.left.text = (
        "Semita Fahrpraxis Manager"
    )

    worksheet.oddFooter.right.text = (
        f"Fahrpraxis {year}"
    )

    # =================================================
    # EXCEL ERSTELLEN
    # =================================================

    buffer = BytesIO()

    workbook.save(
        buffer
    )

    buffer.seek(0)

    # =================================================
    # DATEINAME
    # =================================================

    filename = (

        f"Fahrpraxis_"

        f"{profile.user.last_name}_"

        f"{year}.xlsx"

    )

    # =================================================
    # DOWNLOAD
    # =================================================

    return FileResponse(

        buffer,

        as_attachment=True,

        filename=filename,

        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )

    )
